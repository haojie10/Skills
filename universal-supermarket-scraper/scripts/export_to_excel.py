import json
import os
import sys
import logging
import re
from datetime import datetime

# 配置日志
logging.basicConfig(level=logging.INFO, format='%(message)s')
logger = logging.getLogger(__name__)

# 自动安装依赖
try:
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    logger.info("正在安装 openpyxl...")
    os.system(f"{sys.executable} -m pip install openpyxl -q")
    from openpyxl import Workbook, load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

# 表头定义
HEADERS = ["产品关键词", "产品名称", "产品型号", "产品图片", "产品描述", "产品售价", "产品链接", "超市名称", "国家", "爬取日期"]
FIELD_KEYS = ["keyword", "name", "model", "image_url", "description", "price", "url", "supermarket", "country", "scrape_date"]


def normalize_url(url: str) -> str:
    """
    规范化 URL，剔除可能导致 404 的动态参数或变体 ID。
    特别针对 B&M：移除 ?variant= 和 # 锚点。
    """
    if not url:
        return ""
    
    # 移除查询参数和锚点
    clean_url = url.split('?')[0].split('#')[0].strip()
    
    # 针对 bmstores.co.uk 的特定处理：统一斜杠结尾
    if "bmstores.co.uk" in clean_url and not clean_url.endswith('/'):
        if not re.search(r'\.[a-z0-9]+$', clean_url, re.I):
            clean_url += '/'
            
    return clean_url


def export_products_to_excel(products: list[dict], output_path: str, master_json_path: str = None) -> str:
    """
    将商品数据导出为 Excel，并同步 Master JSON（母表）。
    """
    master_data = []
    existing_urls = set()
    
    # 1. 加载母表并规范化
    if master_json_path and os.path.exists(master_json_path):
        try:
            with open(master_json_path, "r", encoding="utf-8") as f:
                master_data = json.load(f)
            # 对既有数据进行 URL 修复
            for p in master_data:
                if p.get("url"):
                    p["url"] = normalize_url(p["url"])
            existing_urls = {p.get("url") for p in master_data if p.get("url")}
            logger.info(f"已加载母表: {len(master_data)} 条记录")
        except Exception as e:
            logger.warning(f"加载母表失败: {e}")

    current_date = datetime.now().strftime("%Y-%m-%d")

    # 2. 合并新数据并去重
    new_items_to_sync = []
    for p in products:
        raw_url = p.get("url")
        if not raw_url: continue
            
        url = normalize_url(raw_url)
        p["url"] = url 
        
        if url not in existing_urls:
            p["scrape_date"] = p.get("scrape_date") or current_date
            master_data.append(p)
            new_items_to_sync.append(p)
            existing_urls.add(url)

    # 保存母表
    if master_json_path and new_items_to_sync:
        os.makedirs(os.path.dirname(master_json_path) or ".", exist_ok=True)
        with open(master_json_path, "w", encoding="utf-8") as f:
            json.dump(master_data, f, ensure_ascii=False, indent=2)
        logger.info(f"母表已同步，新增 {len(new_items_to_sync)} 条。")

    # 3. 导出到 Excel
    is_new = not os.path.exists(output_path)
    if is_new:
        wb = Workbook()
        ws = wb.active
        ws.title = "商品数据"
        excel_existing_urls = set()
    else:
        try:
            wb = load_workbook(output_path)
            ws = wb.active
            excel_existing_urls = set()
            url_col_idx = FIELD_KEYS.index("url") + 1
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=url_col_idx)
                target = cell.hyperlink.target if cell.hyperlink else cell.value
                if target: excel_existing_urls.add(normalize_url(str(target)))
        except Exception as e:
            logger.warning(f"加载 Excel 失败，重新创建: {e}")
            wb, ws, is_new = Workbook(), Workbook().active, True
            excel_existing_urls = set()

    # 样式
    header_font = Font(name="微软雅黑", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    content_font = Font(name="微软雅黑", size=10)
    link_font = Font(name="微软雅黑", size=10, color="0563C1", underline="single")
    thin_border = Border(left=Side(style="thin"), right=Side(style="thin"), top=Side(style="thin"), bottom=Side(style="thin"))

    if is_new or ws.max_row == 0:
        for col, h in enumerate(HEADERS, 1):
            c = ws.cell(row=1, column=col, value=h)
            c.font, c.fill, c.alignment, c.border = header_font, header_fill, header_alignment, thin_border
        start_row = 2
    else:
        start_row = ws.max_row + 1

    # 只写入 Excel 中没有的数据
    items_to_write = [p for p in master_data if normalize_url(p.get("url")) not in excel_existing_urls]
    
    for i, product in enumerate(items_to_write):
        r = start_row + i
        for c, key in enumerate(FIELD_KEYS, 1):
            val = product.get(key, "")
            cell = ws.cell(row=r, column=c)
            cell.border = thin_border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if key in ["image_url", "url"] and val:
                cell.value, cell.hyperlink, cell.font = "点击查看" if key=="image_url" else "查看商品", val, link_font
            else:
                cell.value, cell.font = val, content_font
        ws.row_dimensions[r].height = 60

    if is_new:
        widths = [15, 40, 15, 12, 50, 12, 12, 15, 10, 15]
        for col, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(col)].width = w
        ws.freeze_panes = "A2"

    if items_to_write:
        os.makedirs(os.path.dirname(output_path) or ".", exist_ok=True)
        wb.save(output_path)
        logger.info(f"Excel 已更新，新增 {len(items_to_write)} 条。")
    else:
        logger.info("Excel 无需更新。")
        
    return output_path


if __name__ == "__main__":
    import argparse
    parser = argparse.ArgumentParser()
    parser.add_argument("json_input")
    parser.add_argument("excel_output")
    parser.add_argument("--master")
    parser.add_argument("--cleanup", action="store_true")
    args = parser.parse_args()

    if not os.path.exists(args.json_input):
        logger.error("输入文件不存在")
        sys.exit(1)

    with open(args.json_input, "r", encoding="utf-8") as f:
        data = json.load(f)
    
    export_products_to_excel(data, args.excel_output, args.master)
    
    if args.cleanup:
        try:
            os.remove(args.json_input)
            logger.info("临时文件已清理")
        except: pass
